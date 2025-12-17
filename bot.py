import os
import time
import sqlite3
from collections import defaultdict
from telegram import (
Update,
InlineKeyboardButton,
InlineKeyboardMarkup,
BotCommand,
BotCommandScopeChat,
)
from telegram.ext import (
ApplicationBuilder,
CommandHandler,
MessageHandler,
CallbackQueryHandler,
ContextTypes,
filters,
)

#================= CONFIG =================

BOT_TOKEN = os.getenv("BOT_TOKEN")
OWNER_ID = 6729390752

UPLOAD_DIR = "/tmp/files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

START_PHOTO_URL = "https://telegram.org/img/t_logo.png"
TRIAL_DURATION = 24 * 3600

BUY_MSG = (
"❌ <b>Access Locked</b>\n\n"
"💳 Buy access from 👉 <b>@indiawsagent</b>"
)

USER_STATE = defaultdict(str)
USER_DATA = defaultdict(dict)

#================= XLSX =================

try:
from openpyxl import load_workbook
XLSX_ENABLED = True
except Exception:
XLSX_ENABLED = False

================= SQLITE =================

conn = sqlite3.connect("data.db", check_same_thread=False)
cur = conn.cursor()

cur.execute("""
CREATE TABLE IF NOT EXISTS sudo_users (
user_id INTEGER PRIMARY KEY,
expires_at INTEGER
)
""")

cur.execute("""
CREATE TABLE IF NOT EXISTS trial_users (
user_id INTEGER PRIMARY KEY,
start_time INTEGER
)
""")
conn.commit()

#================= ACCESS =================

def has_access(uid: int) -> bool:
if uid == OWNER_ID:
return True

now = int(time.time())  

row = cur.execute(  
    "SELECT expires_at FROM sudo_users WHERE user_id=?",  
    (uid,)  
).fetchone()  
if row and row[0] > now:  
    return True  

row = cur.execute(  
    "SELECT start_time FROM trial_users WHERE user_id=?",  
    (uid,)  
).fetchone()  
if row and now - row[0] <= TRIAL_DURATION:  
    return True  

return False

async def ensure_trial(update: Update):
uid = update.effective_user.id
now = int(time.time())

row = cur.execute(  
    "SELECT start_time FROM trial_users WHERE user_id=?",  
    (uid,)  
).fetchone()  

if not row:  
    cur.execute(  
        "INSERT INTO trial_users VALUES (?, ?)",  
        (uid, now)  
    )  
    conn.commit()  

    await update.message.reply_text(  
        "🎁 <b>24 Hours Free Trial Activated</b>",  
        parse_mode="HTML"  
    )

async def deny(update: Update):
await update.message.reply_text(BUY_MSG, parse_mode="HTML")

#================= ERROR =================

async def error_handler(update, context):
print("ERROR:", context.error)
if update and update.effective_message:
try:
await update.effective_message.reply_text(
"⚠️ Something went wrong. Please try again."
)
except:
pass

#================= START / HELP / STATUS =================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
await update.message.reply_photo(
photo=START_PHOTO_URL,
caption=(
"🤖 <b>Smart CV Bot</b>\n\n"
"📇 TXT ⇄ VCF\n"
"📊 XLSX ➜ VCF\n"
"✏️ Rename Contacts (VCF)\n"
"📝 Rename Files\n"
"🎁 24h Free Trial\n\n"
"👉 Use the menu below"
),
parse_mode="HTML"
)

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
await update.message.reply_text(
"/cv_txt_to_vcf\n"
"/cv_vcf_to_txt\n"
"/cv_xlsx_to_vcf\n"
"/renamectc\n"
"/renamefile\n"
"/status"
)

async def status_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
uid = update.effective_user.id
now = int(time.time())

if uid == OWNER_ID:  
    await update.message.reply_text(  
        "👑 <b>Owner</b>\nAccess: Unlimited",  
        parse_mode="HTML"  
    )  
    return  

row = cur.execute(  
    "SELECT expires_at FROM sudo_users WHERE user_id=?",  
    (uid,)  
).fetchone()  
if row and row[0] > now:  
    rem = row[0] - now  
    await update.message.reply_text(  
        f"🧑‍💼 <b>Sudo Access</b>\n"  
        f"⏳ {rem//3600}h {(rem%3600)//60}m left",  
        parse_mode="HTML"  
    )  
    return  

row = cur.execute(  
    "SELECT start_time FROM trial_users WHERE user_id=?",  
    (uid,)  
).fetchone()  
if row and now - row[0] <= TRIAL_DURATION:  
    rem = TRIAL_DURATION - (now - row[0])  
    await update.message.reply_text(  
        f"🎁 <b>Free Trial</b>\n"  
        f"⏳ {rem//3600}h {(rem%3600)//60}m left",  
        parse_mode="HTML"  
    )  
    return  

await update.message.reply_text(BUY_MSG, parse_mode="HTML")

#================= SUDO =================

async def addsudo(update: Update, context: ContextTypes.DEFAULT_TYPE):
if update.effective_user.id != OWNER_ID:
return

if len(context.args) != 2:  
    return await update.message.reply_text(  
        "Usage:\n/addsudo <user_id> <hours>"  
    )  

uid = int(context.args[0])  
hours = int(context.args[1])  
expiry = int(time.time()) + hours * 3600  

cur.execute(  
    "INSERT OR REPLACE INTO sudo_users VALUES (?, ?)",  
    (uid, expiry)  
)  
conn.commit()  

await update.message.reply_text("✅ Sudo granted")  

try:  
    await context.bot.send_message(  
        uid,  
        f"🎉 <b>Sudo Access Granted</b>\n"  
        f"⏳ Valid for {hours} hours",  
        parse_mode="HTML"  
    )  
except:  
    pass

async def delsudo(update: Update, context: ContextTypes.DEFAULT_TYPE):
if update.effective_user.id != OWNER_ID:
return
if not context.args:
return

uid = int(context.args[0])  
cur.execute("DELETE FROM sudo_users WHERE user_id=?", (uid,))  
conn.commit()  
await update.message.reply_text("❌ Sudo removed")

async def listsudo(update: Update, context: ContextTypes.DEFAULT_TYPE):
if update.effective_user.id != OWNER_ID:
return

rows = cur.execute(  
    "SELECT user_id, expires_at FROM sudo_users"  
).fetchall()  

if not rows:  
    return await update.message.reply_text("No sudo users")  

msg = "👑 <b>Sudo Users</b>\n\n"  
now = int(time.time())  

for uid, exp in rows:  
    if exp > now:  
        rem = exp - now  
        msg += f"{uid} → {rem//3600}h {(rem%3600)//60}m\n"  

await update.message.reply_text(msg, parse_mode="HTML")

#================= CORE FUNCTIONS =================

def txt_to_vcf(src, out, name):
with open(src) as f, open(out, "w") as o:
i = 1
for line in f:
num = line.strip()
if num:
o.write(
"BEGIN:VCARD\nVERSION:3.0\n"
f"N:{name}{i};;;\n"
f"FN:{name}{i}\n"
f"TEL:{num}\n"
"END:VCARD\n"
)
i += 1

def vcf_to_txt(src, out):
nums = set()
with open(src, errors="ignore") as f:
for l in f:
if l.startswith("TEL"):
nums.add(l.split(":")[-1].strip())
with open(out, "w") as o:
for n in sorted(nums):
o.write(n + "\n")

def rename_vcf_contacts(src, out, base):
c = 1
with open(src, errors="ignore") as f, open(out, "w") as o:
for line in f:
if line.startswith("FN:"):
o.write(f"FN:{base}{c}\n")
elif line.startswith("N:"):
o.write(f"N:{base}{c};;;;\n")
c += 1
else:
o.write(line)

def xlsx_to_vcf(src, out, name):
wb = load_workbook(src)
sh = wb.active
nums = []
for row in sh.iter_rows(values_only=True):
for cell in row:
if cell and str(cell).isdigit():
nums.append(str(cell))
with open(out, "w") as o:
for i, n in enumerate(nums, 1):
o.write(
"BEGIN:VCARD\nVERSION:3.0\n"
f"N:{name}{i};;;\n"
f"FN:{name}{i}\n"
f"TEL:{n}\n"
"END:VCARD\n"
)

#================= COMMANDS =================

async def cv_txt_to_vcf(update: Update, context: ContextTypes.DEFAULT_TYPE):
await ensure_trial(update)
if not has_access(update.effective_user.id):
return await deny(update)

uid = update.effective_user.id  
USER_STATE[uid] = "TXT"  
USER_DATA[uid] = {"files": []}  

await update.message.reply_text(  
    "📄 Send TXT files",  
    reply_markup=InlineKeyboardMarkup([  
        [InlineKeyboardButton("✅ Done", callback_data="txt_done"),  
         InlineKeyboardButton("❌ Cancel", callback_data="cancel")]  
    ])  
)

async def cv_vcf_to_txt(update: Update, context: ContextTypes.DEFAULT_TYPE):
await ensure_trial(update)
if not has_access(update.effective_user.id):
return await deny(update)

uid = update.effective_user.id  
USER_STATE[uid] = "VCF"  
USER_DATA[uid] = {"files": []}  

await update.message.reply_text(  
    "📇 Send VCF files",  
    reply_markup=InlineKeyboardMarkup([  
        [InlineKeyboardButton("✅ Done", callback_data="vcf_done"),  
         InlineKeyboardButton("❌ Cancel", callback_data="cancel")]  
    ])  
)

async def cv_xlsx_to_vcf(update: Update, context: ContextTypes.DEFAULT_TYPE):
if not XLSX_ENABLED:
return await update.message.reply_text("❌ XLSX support not installed")

await ensure_trial(update)  
if not has_access(update.effective_user.id):  
    return await deny(update)  

uid = update.effective_user.id  
USER_STATE[uid] = "XLSX_FILE"  
USER_DATA[uid] = {}  

await update.message.reply_text("📊 Upload XLSX file")

async def renamectc(update: Update, context: ContextTypes.DEFAULT_TYPE):
await ensure_trial(update)
if not has_access(update.effective_user.id):
return await deny(update)

uid = update.effective_user.id  
USER_STATE[uid] = "RENAMECTC_FILE"  
USER_DATA[uid] = {}  

await update.message.reply_text("📇 Upload VCF file")

async def renamefile(update: Update, context: ContextTypes.DEFAULT_TYPE):
await ensure_trial(update)
if not has_access(update.effective_user.id):
return await deny(update)

uid = update.effective_user.id  
USER_STATE[uid] = "RENAMEFILE_FILE"  
USER_DATA[uid] = {}  

await update.message.reply_text("📝 Upload file")

#================= FILE HANDLER =================

async def file_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
uid = update.effective_user.id
if not update.message.document:
return

doc = update.message.document  
path = os.path.join(UPLOAD_DIR, doc.file_name)  
await (await doc.get_file()).download_to_drive(path)  

st = USER_STATE.get(uid)  

if st in ("TXT", "VCF"):  
    USER_DATA[uid]["files"].append(path)  
    await update.message.reply_text(f"Added: {doc.file_name}")  

elif st == "XLSX_FILE":  
    USER_DATA[uid]["xlsx"] = path  
    USER_STATE[uid] = "XLSX_NAME"  
    await update.message.reply_text("Enter contact name")  

elif st == "RENAMECTC_FILE":  
    USER_DATA[uid]["vcf"] = path  
    USER_STATE[uid] = "RENAMECTC_NAME"  
    await update.message.reply_text("Enter new contact name")  

elif st == "RENAMEFILE_FILE":  
    USER_DATA[uid]["file"] = path  
    USER_DATA[uid]["ext"] = os.path.splitext(doc.file_name)[1]  
    USER_STATE[uid] = "RENAMEFILE_NAME"  
    await update.message.reply_text("Enter new file name")

#================= CALLBACKS =================

async def done_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
q = update.callback_query
uid = q.from_user.id
await q.answer()

files = USER_DATA.get(uid, {}).get("files", [])  
if not files:  
    return await q.message.reply_text("No files uploaded")  

USER_DATA[uid]["type"] = q.data.split("_")[0]  
USER_STATE[uid] = "CONTACT"  
await q.message.reply_text("Enter contact name")

async def cancel_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
q = update.callback_query
uid = q.from_user.id
await q.answer()
USER_STATE.pop(uid, None)
USER_DATA.pop(uid, None)
await q.message.reply_text("Cancelled")

#================= TEXT HANDLER =================

async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
uid = update.effective_user.id
txt = update.message.text.strip()
st = USER_STATE.get(uid)

if st == "CONTACT":  
    for src in USER_DATA[uid]["files"]:  
        out = os.path.join(  
            UPLOAD_DIR,  
            os.path.splitext(os.path.basename(src))[0] +  
            (".vcf" if USER_DATA[uid]["type"] == "txt" else ".txt")  
        )  
        if USER_DATA[uid]["type"] == "txt":  
            txt_to_vcf(src, out, txt)  
        else:  
            vcf_to_txt(src, out)  
        await update.message.reply_document(open(out, "rb"))  

    USER_STATE.pop(uid, None)  
    USER_DATA.pop(uid, None)  

elif st == "XLSX_NAME":  
    out = os.path.join(UPLOAD_DIR, f"{txt}.vcf")  
    xlsx_to_vcf(USER_DATA[uid]["xlsx"], out, txt)  
    await update.message.reply_document(open(out, "rb"))  
    USER_STATE.pop(uid, None)  
    USER_DATA.pop(uid, None)  

elif st == "RENAMECTC_NAME":  
    src = USER_DATA[uid]["vcf"]  
    original_name = os.path.basename(src)  
    out = os.path.join(UPLOAD_DIR, original_name)  

    rename_vcf_contacts(src, out, txt)  

    await update.message.reply_document(  
        open(out, "rb"),  
        filename=original_name  
    )  

    USER_STATE.pop(uid, None)  
    USER_DATA.pop(uid, None)  

elif st == "RENAMEFILE_NAME":  
    out = os.path.join(UPLOAD_DIR, txt + USER_DATA[uid]["ext"])  
    os.rename(USER_DATA[uid]["file"], out)  
    await update.message.reply_document(open(out, "rb"))  
    USER_STATE.pop(uid, None)  
    USER_DATA.pop(uid, None)
#================= MENU =================

async def set_commands(app):
public = [
BotCommand("start", "Start"),
BotCommand("help", "Help"),
BotCommand("status", "Check access"),
BotCommand("cv_txt_to_vcf", "TXT ➜ VCF"),
BotCommand("cv_vcf_to_txt", "VCF ➜ TXT"),
BotCommand("cv_xlsx_to_vcf", "XLSX ➜ VCF"),
BotCommand("renamectc", "Rename contacts"),
BotCommand("renamefile", "Rename file"),
]
await app.bot.set_my_commands(public)

owner = public + [  
    BotCommand("addsudo", "Add sudo"),  
    BotCommand("delsudo", "Delete sudo"),  
    BotCommand("listsudo", "List sudo"),  
]  
await app.bot.set_my_commands(  
    owner,  
    scope=BotCommandScopeChat(chat_id=OWNER_ID)  
)

#================= MAIN =================

def main():
app = ApplicationBuilder().token(BOT_TOKEN).build()

app.add_handler(CommandHandler("start", start))  
app.add_handler(CommandHandler("help", help_cmd))  
app.add_handler(CommandHandler("status", status_cmd))  

app.add_handler(CommandHandler("addsudo", addsudo))  
app.add_handler(CommandHandler("delsudo", delsudo))  
app.add_handler(CommandHandler("listsudo", listsudo))  

app.add_handler(CommandHandler("cv_txt_to_vcf", cv_txt_to_vcf))  
app.add_handler(CommandHandler("cv_vcf_to_txt", cv_vcf_to_txt))  
app.add_handler(CommandHandler("cv_xlsx_to_vcf", cv_xlsx_to_vcf))  

app.add_handler(CommandHandler("renamectc", renamectc))  
app.add_handler(CommandHandler("renamefile", renamefile))  

app.add_handler(CallbackQueryHandler(done_handler, pattern="^(txt|vcf)_done$"))  
app.add_handler(CallbackQueryHandler(cancel_handler, pattern="^cancel$"))  

app.add_handler(MessageHandler(filters.Document.ALL, file_handler))  
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))  

app.add_error_handler(error_handler)  
app.post_init = set_commands  
app.run_polling()

if name == "main":
main()
